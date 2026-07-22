from flask import Flask, render_template, request, redirect, send_file, session
import sqlite3
from openpyxl import Workbook
import csv
import joblib
app = Flask(__name__)
app.secret_key = "sales_dashboard_secret"

model = joblib.load("sales_model.pkl")


# ---------------- HOME ----------------

@app.route("/", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        username = request.form["username"]
        password = request.form["password"]

        if username == "admin" and password == "admin123":

            session["user"] = username

            return redirect("/dashboard")

        else:
            return render_template(
                "login.html",
                error="Invalid Username or Password"
            )

    return render_template("login.html")


# ---------------- DASHBOARD ----------------

@app.route("/dashboard")
def dashboard():
    if "user" not in session:
        return redirect("/")
    conn = sqlite3.connect("database/sales.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    search = request.args.get("search", "")

    if search:
        cur.execute(
            "SELECT * FROM companies WHERE name LIKE ?",
            ('%' + search + '%',)
        )
        companies = cur.fetchall()
    else:
        cur.execute("SELECT * FROM companies")
        companies = cur.fetchall()

    cur.execute("SELECT name, revenue FROM companies ORDER BY revenue DESC LIMIT 1")
    top_revenue = cur.fetchone()

    cur.execute("SELECT name, profit FROM companies ORDER BY profit DESC LIMIT 1")
    top_profit = cur.fetchone()

    cur.execute("SELECT name, orders FROM companies ORDER BY orders DESC LIMIT 1")
    top_orders = cur.fetchone()

    cur.execute("SELECT COUNT(*) FROM companies")
    total_companies = cur.fetchone()[0]

    cur.execute("SELECT SUM(revenue) FROM companies")
    total_revenue = cur.fetchone()[0] or 0

    cur.execute("SELECT SUM(profit) FROM companies")
    total_profit = cur.fetchone()[0] or 0

    cur.execute("SELECT SUM(orders) FROM companies")
    total_orders = cur.fetchone()[0] or 0
    # Hardcoded values for now
    cur.execute("SELECT COUNT(*) FROM customers")
    total_customers = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM products")
    total_products = cur.fetchone()[0]

    conn.close()

    return render_template(
       "dashboard.html",
       companies=companies,
       total_companies=total_companies,
       total_customers=total_customers,
       total_products=total_products,
       total_revenue=total_revenue,
       total_profit=total_profit,
       total_orders=total_orders,
       top_revenue=top_revenue,
       top_profit=top_profit,
       top_orders=top_orders,
       search=search
)

# ---------------- ADD COMPANY ----------------

@app.route("/add_company", methods=["GET", "POST"])
def add_company():

    if request.method == "POST":

        name = request.form["name"]
        revenue = request.form["revenue"]
        profit = request.form["profit"]
        orders = request.form["orders"]

        conn = sqlite3.connect("database/sales.db")
        cur = conn.cursor()

        cur.execute(
            """
            INSERT INTO companies
            (name,revenue,profit,orders)
            VALUES (?,?,?,?)
            """,
            (name, revenue, profit, orders)
        )

        conn.commit()
        conn.close()

        return redirect("/dashboard")

    return render_template("add_company.html")


# ---------------- EDIT COMPANY ----------------

@app.route("/edit/<int:id>", methods=["GET", "POST"])
def edit(id):

    conn = sqlite3.connect("database/sales.db")
    conn.row_factory = sqlite3.Row

    cur = conn.cursor()

    if request.method == "POST":

        name = request.form["name"]
        revenue = request.form["revenue"]
        profit = request.form["profit"]
        orders = request.form["orders"]

        cur.execute(
            """
            UPDATE companies
            SET
            name=?,
            revenue=?,
            profit=?,
            orders=?
            WHERE id=?
            """,
            (name, revenue, profit, orders, id)
        )

        conn.commit()
        conn.close()

        return redirect("/dashboard")

    cur.execute("SELECT * FROM companies WHERE id=?", (id,))
    company = cur.fetchone()

    conn.close()

    return render_template(
        "edit_company.html",
        company=company
    )


# ---------------- DELETE COMPANY ----------------

@app.route("/delete/<int:id>")
def delete(id):

    conn = sqlite3.connect("database/sales.db")
    cur = conn.cursor()

    cur.execute(
        "DELETE FROM companies WHERE id=?",
        (id,)
    )

    conn.commit()
    conn.close()

    return redirect("/dashboard")


# ---------------- ANALYTICS ----------------

@app.route("/analytics")
def analytics():

    conn = sqlite3.connect("database/sales.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("SELECT * FROM companies")

    rows = cur.fetchall()

    companies = []

    for row in rows:
        companies.append({
            "id": row["id"],
            "name": row["name"],
            "revenue": float(row["revenue"]),
            "profit": float(row["profit"]),
            "orders": int(row["orders"])
        })

    conn.close()

    return render_template(
        "analytics.html",
        companies=companies
    )

# ---------------- PRODUCTS ----------------

@app.route("/products")
def products():

    conn = sqlite3.connect("database/sales.db")
    conn.row_factory = sqlite3.Row

    cur = conn.cursor()
    search = request.args.get("search", "")

    if search:
        cur.execute(
            "SELECT * FROM products WHERE name LIKE ?",
        ('%' + search + '%',)
        )
    else:
        cur.execute("SELECT * FROM products")

    products = cur.fetchall()


    conn.close()

    return render_template(
        "products.html",
        products=products
    )
@app.route("/add_product", methods=["GET", "POST"])
def add_product():

    if request.method == "POST":

        name = request.form["name"]
        brand = request.form["brand"]
        price = request.form["price"]
        stock = request.form["stock"]

        conn = sqlite3.connect("database/sales.db")
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO products
            (name, brand, price, stock)
            VALUES (?, ?, ?, ?)
        """, (name, brand, price, stock))

        conn.commit()
        conn.close()

        return redirect("/products")

    return render_template("add_product.html")

# ---------------- EDIT PRODUCT ----------------

@app.route("/edit_product/<int:id>", methods=["GET", "POST"])
def edit_product(id):

    conn = sqlite3.connect("database/sales.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    if request.method == "POST":

        name = request.form["name"]
        brand = request.form["brand"]
        price = request.form["price"]
        stock = request.form["stock"]

        cur.execute("""
            UPDATE products
            SET
                name=?,
                brand=?,
                price=?,
                stock=?
            WHERE id=?
        """, (name, brand, price, stock, id))

        conn.commit()
        conn.close()

        return redirect("/products")

    cur.execute("SELECT * FROM products WHERE id=?", (id,))
    product = cur.fetchone()

    conn.close()

    return render_template(
        "edit_product.html",
        product=product
    )

# ---------------- DELETE PRODUCT ----------------

@app.route("/delete_product/<int:id>")
def delete_product(id):

    conn = sqlite3.connect("database/sales.db")
    cur = conn.cursor()

    cur.execute(
        "DELETE FROM products WHERE id=?",
        (id,)
    )

    conn.commit()
    conn.close()

    return redirect("/products")

# ---------------- COMPANY DETAILS ----------------

@app.route("/company/<int:id>")
def company_details(id):

    conn = sqlite3.connect("database/sales.db")
    conn.row_factory = sqlite3.Row

    cur = conn.cursor()

    cur.execute(
        "SELECT * FROM companies WHERE id=?",
        (id,)
    )

    company = cur.fetchone()

    conn.close()

    return render_template(
        "company_details.html",
        company=company
    )
# ---------------- CUSTOMERS ----------------

@app.route("/customers")
def customers():

    conn = sqlite3.connect("database/sales.db")
    conn.row_factory = sqlite3.Row

    cur = conn.cursor()

    search = request.args.get("search", "")

    if search:
        cur.execute(
            "SELECT * FROM customers WHERE name LIKE ?",
            ('%' + search + '%',)
        )
    else:
        cur.execute("SELECT * FROM customers")

    customers = cur.fetchall()

    conn.close()

    return render_template(
        "customers.html",
        customers=customers
    )


# ---------------- ADD CUSTOMER ----------------

@app.route("/add_customer", methods=["GET", "POST"])
def add_customer():

    if request.method == "POST":

        name = request.form["name"]
        email = request.form["email"]
        phone = request.form["phone"]
        orders = request.form["orders"]

        conn = sqlite3.connect("database/sales.db")
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO customers
            (name, email, phone, orders)
            VALUES (?, ?, ?, ?)
        """, (name, email, phone, orders))

        conn.commit()
        conn.close()

        return redirect("/customers")

    return render_template("add_customer.html")


# ---------------- EDIT CUSTOMER ----------------

@app.route("/edit_customer/<int:id>", methods=["GET", "POST"])
def edit_customer(id):

    conn = sqlite3.connect("database/sales.db")
    conn.row_factory = sqlite3.Row

    cur = conn.cursor()

    if request.method == "POST":

        name = request.form["name"]
        email = request.form["email"]
        phone = request.form["phone"]
        orders = request.form["orders"]

        cur.execute("""
            UPDATE customers
            SET
                name=?,
                email=?,
                phone=?,
                orders=?
            WHERE id=?
        """, (name, email, phone, orders, id))

        conn.commit()
        conn.close()

        return redirect("/customers")

    cur.execute(
        "SELECT * FROM customers WHERE id=?",
        (id,)
    )

    customer = cur.fetchone()

    conn.close()

    return render_template(
        "edit_customer.html",
        customer=customer
    )


# ---------------- DELETE CUSTOMER ----------------

@app.route("/delete_customer/<int:id>")
def delete_customer(id):

    conn = sqlite3.connect("database/sales.db")
    cur = conn.cursor()

    cur.execute(
        "DELETE FROM customers WHERE id=?",
        (id,)
    )

    conn.commit()
    conn.close()

    return redirect("/customers")

# ---------------- RUN ----------------
# ---------------- ORDERS ----------------

@app.route("/orders")
def orders():

    conn = sqlite3.connect("database/sales.db")
    conn.row_factory = sqlite3.Row

    cur = conn.cursor()

    cur.execute("SELECT * FROM orders")

    orders = cur.fetchall()

    conn.close()

    return render_template(
        "orders.html",
        orders=orders
    )

@app.route("/add_order", methods=["GET", "POST"])
def add_order():

    if request.method == "POST":

        customer = request.form["customer"]
        product = request.form["product"]
        quantity = request.form["quantity"]
        price = request.form["price"]
        status = request.form["status"]

        conn = sqlite3.connect("database/sales.db")
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO orders
            (customer, product, quantity, price, status)
            VALUES (?, ?, ?, ?, ?)
        """, (customer, product, quantity, price, status))

        conn.commit()
        conn.close()

        return redirect("/orders")

    return render_template("add_order.html")


@app.route("/delete_order/<int:id>")
def delete_order(id):

    conn = sqlite3.connect("database/sales.db")
    cur = conn.cursor()

    cur.execute(
        "DELETE FROM orders WHERE id=?",
        (id,)
    )

    conn.commit()
    conn.close()

    return redirect("/orders")

@app.route("/ai_prediction")
def ai_prediction():

    conn = sqlite3.connect("database/sales.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("SELECT * FROM companies")
    companies = cur.fetchall()

    predictions = []

    for company in companies:

        revenue = float(company["revenue"])
        profit = float(company["profit"])
        current_orders = int(company["orders"])

        # Random Forest prediction
        predicted_orders = int(model.predict([[revenue, profit]])[0])

        predicted_revenue = revenue * 1.10
        predicted_profit = profit * 1.08

        if predicted_orders >= current_orders + 20:
            recommendation = "🟢 High Growth Expected"

        elif predicted_orders >= current_orders:
            recommendation = "🟡 Stable Growth"

        else:
            recommendation = "🔴 Sales May Decrease"

        predictions.append({

            "name": company["name"],

            "revenue": revenue,

            "profit": profit,

            "orders": current_orders,

            "predicted_revenue": round(predicted_revenue,2),

            "predicted_profit": round(predicted_profit,2),

            "predicted_orders": predicted_orders,

            "recommendation": recommendation

        })

    conn.close()

    return render_template(
        "ai_prediction.html",
        predictions=predictions
    )
# ---------------- EXPORT COMPANIES ----------------

@app.route("/export_companies")
def export_companies():

    conn = sqlite3.connect("database/sales.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("SELECT * FROM companies")
    companies = cur.fetchall()

    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Companies"

    # Headings
    ws.append([
        "ID",
        "Company",
        "Revenue",
        "Profit",
        "Orders"
    ])

    # Data
    for company in companies:
        ws.append([
            company["id"],
            company["name"],
            company["revenue"],
            company["profit"],
            company["orders"]
        ])

    filename = "companies.xlsx"
    wb.save(filename)

    return send_file(
        filename,
        as_attachment=True
    )
@app.route("/export_dashboard")
def export_dashboard():

    conn = sqlite3.connect("database/sales.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    wb = Workbook()

    # ---------------- Summary Sheet ----------------
    ws = wb.active
    ws.title = "Summary"

    cur.execute("SELECT COUNT(*) FROM companies")
    total_companies = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM products")
    total_products = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM customers")
    total_customers = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM orders")
    total_orders = cur.fetchone()[0]

    cur.execute("SELECT SUM(revenue) FROM companies")
    total_revenue = cur.fetchone()[0] or 0

    cur.execute("SELECT SUM(profit) FROM companies")
    total_profit = cur.fetchone()[0] or 0
    # Top Revenue Company
    cur.execute("SELECT name, revenue FROM companies ORDER BY revenue DESC LIMIT 1")
    top_revenue = cur.fetchone()

    # Top Profit Company
    cur.execute("SELECT name, profit FROM companies ORDER BY profit DESC LIMIT 1")
    top_profit = cur.fetchone()

    # Top Orders Company
    cur.execute("SELECT name, orders FROM companies ORDER BY orders DESC LIMIT 1")
    top_orders = cur.fetchone()

    ws.append(["Metric", "Value"])
    ws.append(["Total Companies", total_companies])
    ws.append(["Total Products", total_products])
    ws.append(["Total Customers", total_customers])
    ws.append(["Total Orders", total_orders])
    ws.append(["Total Revenue", total_revenue])
    ws.append(["Total Profit", total_profit])

    # ---------------- Companies Sheet ----------------
    company_sheet = wb.create_sheet("Companies")
    company_sheet.append(["ID", "Name", "Revenue", "Profit", "Orders"])

    cur.execute("SELECT * FROM companies")
    for row in cur.fetchall():
        company_sheet.append([
            row["id"],
            row["name"],
            row["revenue"],
            row["profit"],
            row["orders"]
        ])

    conn.close()

    filename = "Dashboard_Report.xlsx"
    wb.save(filename)

    return send_file(filename, as_attachment=True)
@app.route("/export_products")
def export_products():

    conn = sqlite3.connect("database/sales.db")
    conn.row_factory = sqlite3.Row

    cur = conn.cursor()
    cur.execute("SELECT * FROM products")

    products = cur.fetchall()

    conn.close()

    filename = "products.csv"

    with open(filename, "w", newline="") as file:

        writer = csv.writer(file)

        writer.writerow([
            "ID",
            "Name",
            "Brand",
            "Price",
            "Stock"
        ])

        for product in products:

            writer.writerow([
                product["id"],
                product["name"],
                product["brand"],
                product["price"],
                product["stock"]
            ])

    return send_file(
        filename,
        as_attachment=True
    )
@app.route("/export_customers")
def export_customers():

    conn = sqlite3.connect("database/sales.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("SELECT * FROM customers")
    customers = cur.fetchall()

    conn.close()

    filename = "customers.csv"

    with open(filename, "w", newline="") as file:

        writer = csv.writer(file)

        writer.writerow([
            "ID",
            "Name",
            "Email",
            "Phone",
            "Orders"
        ])

        for customer in customers:
            writer.writerow([
                customer["id"],
                customer["name"],
                customer["email"],
                customer["phone"],
                customer["orders"]
            ])

    return send_file(filename, as_attachment=True)
@app.route("/export_orders")
def export_orders():

    conn = sqlite3.connect("database/sales.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("SELECT * FROM orders")
    orders = cur.fetchall()

    conn.close()

    filename = "orders.csv"

    with open(filename, "w", newline="") as file:

        writer = csv.writer(file)

        writer.writerow([
            "ID",
            "Customer",
            "Product",
            "Quantity",
            "Price",
            "Status"
        ])

        for order in orders:
            writer.writerow([
                order["id"],
                order["customer"],
                order["product"],
                order["quantity"],
                order["price"],
                order["status"]
            ])

    return send_file(filename, as_attachment=True)
@app.route("/logout")
def logout():

    session.pop("user", None)

    return redirect("/")
if __name__ == "__main__":
    app.run(debug=True)